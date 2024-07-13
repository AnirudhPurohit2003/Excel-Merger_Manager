#import python libraries to use them in code;
#Make sure the libraries you want to use are installed in your Python environment. 
# If not, you can install them using pip, Python's package installer. For example:
# used this command in terminal - pip install <library name>.
# Eg. --> pip install pandas

import pandas as pd
import os
from openpyxl import load_workbook

# Folder containing your CSV files
folder_path = 'Enter folder path or folder name (if excle files availabel in same folder)'

# Name of the output Excel file
output_excel_file = '<merged_file.xlsx> (Write file name which is used for save merged file)'

# Create a new Excel writer object
excel_writer = pd.ExcelWriter(output_excel_file, engine='openpyxl')

# Iterate over each CSV file in the folder
for filename in os.listdir(folder_path):
    if filename.endswith('.csv'):
        # Read the CSV file into a DataFrame
        file_path = os.path.join(folder_path, filename)
        df = pd.read_csv(file_path)
        
        # Use the filename (without extension) as the sheet name
        sheet_name = os.path.splitext(filename)[0]
        
        # Truncate the sheet name to 31 characters if necessary
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        # Write the DataFrame to the Excel workbook
        df.to_excel(excel_writer, sheet_name=sheet_name, index=False)

# Save the Excel workbook
excel_writer.close()

# Load the workbook to freeze the first row
workbook = load_workbook(output_excel_file)

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    # Freeze the first row
    sheet.freeze_panes = sheet['A2']

# Save the workbook again
workbook.save(output_excel_file)

print('All CSV files have been merged into', output_excel_file, 'with the first row frozen')